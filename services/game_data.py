import io
import logging
import asyncio
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from services.yandex_sheets import YandexDiskAPI
from services import local_store
from config_reader import config


class GameDataManager:
    """Менеджер для работы с данными игры через Excel файл"""
    
    def __init__(self):
        self.yandex = YandexDiskAPI(config.yadisk_token.get_secret_value())
        self.file_path = config.yadisk_file_path
        self._copy_file_path = self.file_path.replace('.xlsx', '_copy.xlsx')  # Путь к копии
        self._cache: Optional[bytes] = None
        self._cache_time: Optional[datetime] = None
        self._cache_ttl = timedelta(minutes=5)  # Кэш на 5 минут
        self._main_file_mtime: Optional[datetime] = None  # Время последнего изменения основного файла
        # Локальная БД + отложенная синхронизация
        self._sync_task = None
        self._sync_delay_seconds = 60
    
    async def _get_working_file_path(self) -> str:
        """Определяет, с каким файлом работать: основным или копией"""
        try:
            # Проверяем, существует ли основной файл
            try:
                main_info = await self.yandex.get_file_info(self.file_path)
            except:
                # Основного файла нет, создадим новый
                return self.file_path
            
            # Проверяем, существует ли копия
            try:
                copy_info = await self.yandex.get_file_info(self._copy_file_path)
                # Копия существует, проверяем, свежая ли она
                
                # Преобразуем время модификации (формат: "2024-01-01T12:00:00+00:00" или "2024-01-01T12:00:00Z")
                try:
                    copy_mtime_str = copy_info.get('modified', '')
                    main_mtime_str = main_info.get('modified', '')
                    
                    if copy_mtime_str and main_mtime_str:
                        # Парсим ISO формат (Python 3.11+ поддерживает fromisoformat)
                        # Обрабатываем разные форматы
                        if copy_mtime_str.endswith('Z'):
                            copy_mtime_str = copy_mtime_str[:-1] + '+00:00'
                        if main_mtime_str.endswith('Z'):
                            main_mtime_str = main_mtime_str[:-1] + '+00:00'
                        
                        copy_mtime = datetime.fromisoformat(copy_mtime_str)
                        main_mtime = datetime.fromisoformat(main_mtime_str)
                        
                        # Если основной файл новее копии - обновляем копию
                        if main_mtime > copy_mtime:
                            logging.info("Основной файл новее копии, обновляем копию...")
                            await self._refresh_copy()
                    
                    return self._copy_file_path
                except Exception as e:
                    logging.warning(f"Ошибка при сравнении времени модификации: {e}, используем копию")
                    return self._copy_file_path
            except:
                # Копии нет, создаем её из основного файла
                logging.info("Создаем копию файла для работы...")
                await self._create_copy()
                return self._copy_file_path
        except Exception as e:
            # Если не удалось создать копию, работаем с основным файлом
            logging.warning(f"Не удалось создать копию, работаем с основным файлом: {e}")
            return self.file_path
    
    async def _create_copy(self) -> None:
        """Создает копию основного файла"""
        try:
            # Проверяем, существует ли основной файл
            await self.yandex.get_file_info(self.file_path)
            # Копируем файл
            await self.yandex.copy_file(self.file_path, self._copy_file_path)
            logging.info(f"Создана копия файла: {self._copy_file_path}")
        except Exception as e:
            error_str = str(e).lower()
            if "404" in error_str or "not found" in error_str:
                # Основного файла нет, создаем новый и копию
                data = await self._create_new_file()
                await self.yandex.upload_file(data, self._copy_file_path, overwrite=True)
            else:
                raise
    
    async def _refresh_copy(self) -> None:
        """Обновляет копию из основного файла"""
        try:
            # Скачиваем основной файл
            main_data = await self.yandex.download_file(self.file_path)
            # Загружаем как копию
            await self.yandex.upload_file(main_data, self._copy_file_path, overwrite=True)
            logging.info("Копия обновлена из основного файла")
        except Exception as e:
            logging.error(f"Ошибка при обновлении копии: {e}")
            raise
    
    async def _get_file_data(self, force_refresh: bool = False, use_copy: bool = True) -> bytes:
        """Получает данные файла с кэшированием (оптимизированная версия без постоянных проверок копии)"""
        now = datetime.now()
        if not force_refresh and self._cache and self._cache_time:
            if now - self._cache_time < self._cache_ttl:
                return self._cache
        
        try:
            # Сначала пробуем загрузить копию (быстрее)
            try:
                data = await self.yandex.download_file(self._copy_file_path)
                self._cache = data
                self._cache_time = now
                return data
            except:
                # Если копии нет, загружаем основной файл
                data = await self.yandex.download_file(self.file_path)
                # Создаем копию для будущих операций
                try:
                    await self.yandex.upload_file(data, self._copy_file_path, overwrite=True)
                except:
                    pass
                self._cache = data
                self._cache_time = now
                return data
        except Exception as e:
            # Если файла нет, создаем новый
            error_str = str(e).lower()
            if "404" in error_str or "not found" in error_str or "does not exist" in error_str:
                logging.info("Файл не найден на Яндекс.Диске, создаем новый...")
                return await self._create_new_file()
            logging.error(f"Ошибка при загрузке файла с Яндекс.Диска: {e}")
            raise
    
    def _create_empty_data_structure(self) -> Dict[str, Any]:
        """Создает пустую структуру данных"""
        return {
            "participants": [],
            "reports": [],
            "settings": {}
        }
    
    async def _create_new_file(self) -> bytes:
        """Создает новый файл Excel с базовой структурой"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Участники"
        
        # Заголовки
        headers = ["User ID", "Username", "Full Name", "Game Name", "Registered Date", "Status"] + \
                  [f"Goal {i}" for i in range(1, 11)]
        ws.append(headers)
        
        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Лист для отчетов
        ws_reports = wb.create_sheet("Отчеты")
        report_headers = ["User ID", "Day", "Date", "Goal 1", "Goal 2", "Goal 3", "Goal 4", "Goal 5",
                         "Goal 6", "Goal 7", "Goal 8", "Goal 9", "Goal 10", "Rest Day"]
        ws_reports.append(report_headers)
        
        # Стили для заголовков отчетов
        for cell in ws_reports[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Лист для настроек
        ws_settings = wb.create_sheet("Настройки")
        settings_headers = ["Key", "Value"]
        ws_settings.append(settings_headers)
        
        # Стили для заголовков настроек
        for cell in ws_settings[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Сохраняем в байты
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        data = buffer.read()
        buffer.close()
        
        # Сначала сохраняем в основной файл (при создании нового файла)
        await self.yandex.upload_file(data, self.file_path, overwrite=True)
        # Затем создаем копию
        try:
            await self.yandex.copy_file(self.file_path, self._copy_file_path)
        except:
            # Если копирование не удалось, просто загружаем копию отдельно
            await self.yandex.upload_file(data, self._copy_file_path, overwrite=True)
        
        self._cache = data
        self._cache_time = datetime.now()
        
        return data
    
    async def _build_excel_bytes(self, data: Dict[str, Any]) -> bytes:
        """Строит Excel байты из словаря данных."""
        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)
        # Участники
        ws = wb.create_sheet("Участники")
        headers = ["User ID", "Username", "Full Name", "Game Name", "Registered Date", "Status"] + \
                  [f"Goal {i}" for i in range(1, 11)]
        ws.append(headers)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for participant in data["participants"]:
            row = [
                participant["user_id"],
                participant["username"],
                participant["full_name"],
                participant["game_name"],
                participant["registered_date"],
                participant["status"],
            ] + participant["goals"]
            ws.append(row)
        # Отчеты
        ws_reports = wb.create_sheet("Отчеты")
        report_headers = ["User ID", "Day", "Date", "Goal 1", "Goal 2", "Goal 3", "Goal 4", "Goal 5",
                          "Goal 6", "Goal 7", "Goal 8", "Goal 9", "Goal 10", "Rest Day"]
        ws_reports.append(report_headers)
        for cell in ws_reports[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for report in data["reports"]:
            row = [report["user_id"], report["day"], report["date"]] + report["progress"] + ["Да" if report.get("rest_day") else "Нет"]
            ws_reports.append(row)
        # Настройки
        ws_settings = wb.create_sheet("Настройки")
        settings_headers = ["Key", "Value"]
        ws_settings.append(settings_headers)
        for cell in ws_settings[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for key, value in (data.get("settings", {}) or {}).items():
            if value is not None:
                ws_settings.append([key, value])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        out = buffer.read()
        buffer.close()
        return out

    async def _schedule_sync(self) -> None:
        """Планирует отложенную синхронизацию на Я.Диск."""
        async def _job():
            try:
                await asyncio.sleep(self._sync_delay_seconds)
                data = await local_store.get_json("all_data")
                if not data:
                    return
                file_data = await self._build_excel_bytes(data)
                try:
                    await self.yandex.upload_file(file_data, self._copy_file_path, overwrite=True)
                except Exception as e:
                    logging.error(f"Ошибка сохранения копии на Я.Диске: {e}")
                    return
                try:
                    await self.yandex.copy_file(self._copy_file_path, self.file_path)
                except Exception as e:
                    logging.warning(f"Не удалось синхронизировать в основной файл: {e}")
            except Exception as e:
                logging.error(f"Ошибка фоновой синхронизации: {e}")

        # Отменяем предыдущую задачу, если еще не стартовала
        try:
            if self._sync_task and not self._sync_task.done():
                self._sync_task.cancel()
        except Exception:
            pass
        loop = asyncio.get_running_loop()
        self._sync_task = loop.create_task(_job())

    async def get_all_data(self) -> Dict[str, Any]:
        """Получает все данные из локальной БД (или инициализирует из Я.Диска один раз)."""
        try:
            data = await local_store.get_json("all_data")
            if data is not None:
                return data
            # Инициализация из Я.Диска, если локально пусто
            file_data = await self._get_file_data()
            wb = load_workbook(io.BytesIO(file_data))
            
            # Проверяем и создаем лист "Участники", если его нет
            if "Участники" not in wb.sheetnames:
                logging.warning("Лист 'Участники' не найден в файле, возвращаем пустую структуру")
                return self._create_empty_data_structure()
            
            # Участники
            ws = wb["Участники"]
            participants = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:  # User ID не пустой
                    # Проверяем, что достаточно столбцов
                    goals = []
                    if len(row) > 6:
                        goals = [row[5+i] or "" if 5+i < len(row) else "" for i in range(1, 11)]
                    else:
                        goals = [""] * 10
                        
                    participants.append({
                        "user_id": row[0],
                        "username": row[1] if len(row) > 1 else "",
                        "full_name": row[2] if len(row) > 2 else "",
                        "game_name": row[3] if len(row) > 3 else "",
                        "registered_date": row[4] if len(row) > 4 else "",
                        "status": row[5] if len(row) > 5 else "active",
                        "goals": goals
                    })
            
            # Отчеты
            reports = []
            if "Отчеты" in wb.sheetnames:
                ws_reports = wb["Отчеты"]
                for row in ws_reports.iter_rows(min_row=2, values_only=True):
                    if row and row[0] is not None:
                        progress = []
                        if len(row) > 3:
                            progress = [row[2+i] or "" if 2+i < len(row) else "" for i in range(1, 11)]
                        else:
                            progress = [""] * 10
                        
                        reports.append({
                            "user_id": row[0],
                            "day": row[1] if len(row) > 1 else 1,
                            "date": row[2] if len(row) > 2 else "",
                            "progress": progress,
                            "rest_day": row[13] == "Да" if len(row) > 13 and row[13] else False
                        })
            
            # Настройки
            settings = {}
            if "Настройки" in wb.sheetnames:
                ws_settings = wb["Настройки"]
                for row in ws_settings.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        key = str(row[0])
                        value = row[1] if len(row) > 1 else None
                        settings[key] = value
            
            data_dict = {
                "participants": participants,
                "reports": reports,
                "settings": settings
            }
            await local_store.set_json("all_data", data_dict)
            return data_dict
        except Exception as e:
            logging.error(f"Ошибка при чтении данных из файла: {e}")
            # Возвращаем пустую структуру при ошибке
            empty = self._create_empty_data_structure()
            await local_store.set_json("all_data", empty)
            return empty

    async def refresh_local_cache_from_remote(self) -> Dict[str, Any]:
        """Принудительно перечитывает данные из удаленного файла и обновляет локальный кеш."""
        try:
            # Если локальные данные свежее минуты — не перезатираем
            try:
                local_updated_at = await local_store.get_updated_at("all_data")
            except Exception:
                local_updated_at = None
            from time import time
            now_epoch = int(time())
            if local_updated_at and (now_epoch - local_updated_at) < 60:
                current = await local_store.get_json("all_data")
                return current or self._create_empty_data_structure()

            # Смотрим, новее ли удаленный файл локальных данных
            remote_mtime = None
            try:
                info = await self.yandex.get_file_info(self._copy_file_path)
                mtime_str = info.get("modified", "")
                if mtime_str.endswith('Z'):
                    mtime_str = mtime_str[:-1] + '+00:00'
                from datetime import datetime
                remote_dt = datetime.fromisoformat(mtime_str) if mtime_str else None
                if remote_dt:
                    remote_mtime = int(remote_dt.timestamp())
            except Exception:
                remote_mtime = None

            if local_updated_at and remote_mtime and remote_mtime <= local_updated_at:
                current = await local_store.get_json("all_data")
                return current or self._create_empty_data_structure()

            file_data = await self._get_file_data(force_refresh=True)
            wb = load_workbook(io.BytesIO(file_data))

            # Если нет листа участников — сохранить пустую структуру
            if "Участники" not in wb.sheetnames:
                data = self._create_empty_data_structure()
                await local_store.set_json("all_data", data)
                return data

            # Участники
            ws = wb["Участники"]
            participants = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    goals = []
                    if len(row) > 6:
                        goals = [row[5+i] or "" if 5+i < len(row) else "" for i in range(1, 11)]
                    else:
                        goals = [""] * 10
                    participants.append({
                        "user_id": row[0],
                        "username": row[1] if len(row) > 1 else "",
                        "full_name": row[2] if len(row) > 2 else "",
                        "game_name": row[3] if len(row) > 3 else "",
                        "registered_date": row[4] if len(row) > 4 else "",
                        "status": row[5] if len(row) > 5 else "active",
                        "goals": goals
                    })

            # Отчеты
            reports = []
            if "Отчеты" in wb.sheetnames:
                ws_reports = wb["Отчеты"]
                for row in ws_reports.iter_rows(min_row=2, values_only=True):
                    if row and row[0] is not None:
                        progress = []
                        if len(row) > 3:
                            progress = [row[2+i] or "" if 2+i < len(row) else "" for i in range(1, 11)]
                        else:
                            progress = [""] * 10
                        reports.append({
                            "user_id": row[0],
                            "day": row[1] if len(row) > 1 else 1,
                            "date": row[2] if len(row) > 2 else "",
                            "progress": progress,
                            "rest_day": row[13] == "Да" if len(row) > 13 and row[13] else False
                        })

            # Настройки
            settings = {}
            if "Настройки" in wb.sheetnames:
                ws_settings = wb["Настройки"]
                for row in ws_settings.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        key = str(row[0])
                        value = row[1] if len(row) > 1 else None
                        settings[key] = value

            data_dict = {
                "participants": participants,
                "reports": reports,
                "settings": settings
            }
            await local_store.set_json("all_data", data_dict)
            # Инвалидируем in-memory кеш
            self._cache = None
            self._cache_time = None
            return data_dict
        except Exception as e:
            logging.error(f"Ошибка принудительного обновления данных: {e}")
            # В случае ошибки не ломаемся: возвращаем локальные данные
            current = await local_store.get_json("all_data")
            return current or self._create_empty_data_structure()
    
    async def save_data(self, data: Dict[str, Any], sync_to_main: bool = False) -> None:
        """Сохраняет данные локально и планирует синхронизацию на Я.Диск."""
        await local_store.set_json("all_data", data)
        # инвалидация in-memory
        self._cache = None
        self._cache_time = None
        # Планируем фоновой синк; если sync_to_main=True — синкнем раньше (через малую задержку)
        if sync_to_main:
            prev = self._sync_delay_seconds
            self._sync_delay_seconds = 2
            await self._schedule_sync()
            self._sync_delay_seconds = prev
        else:
            await self._schedule_sync()
    
    async def get_settings(self) -> Dict[str, Any]:
        """Получает настройки из файла"""
        data = await self.get_all_data()
        return data.get("settings", {})
    
    async def save_settings(self, settings: Dict[str, Any]) -> None:
        """Сохраняет настройки в файл"""
        data = await self.get_all_data()
        data["settings"] = settings
        await self.save_data(data, sync_to_main=True)
        # Инвалидируем кеш, чтобы изменения применились сразу
        self._cache = None
        self._cache_time = None
    
    async def get_chat_config(self) -> Dict[str, Optional[int]]:
        """Получает конфигурацию чата (chat_id и thread_id)"""
        settings = await self.get_settings()
        return {
            "chat_id": int(settings.get("chat_id", 0)) if settings.get("chat_id") else None,
            "thread_id": int(settings.get("thread_id", 0)) if settings.get("thread_id") else None
        }
    
    async def save_chat_config(self, chat_id: Optional[int], thread_id: Optional[int]) -> None:
        """Сохраняет конфигурацию чата"""
        settings = await self.get_settings()
        if chat_id:
            settings["chat_id"] = str(chat_id)
        if thread_id:
            settings["thread_id"] = str(thread_id)
        await self.save_settings(settings)
        # Инвалидируем кеш (уже делается в save_settings, но для ясности)
        self._cache = None
        self._cache_time = None
    
    def is_user_registered(self, user_id: int, data: Dict) -> bool:
        """Проверяет, зарегистрирован ли пользователь"""
        for participant in data["participants"]:
            if participant["user_id"] == user_id:
                return True
        return False
    
    def register_user(self, user_id: int, username: str, full_name: str, game_name: str, data: Dict) -> None:
        """Регистрирует нового пользователя"""
        if self.is_user_registered(user_id, data):
            return
        
        participant = {
            "user_id": user_id,
            "username": username,
            "full_name": full_name,
            "game_name": game_name,
            "registered_date": datetime.now().strftime("%Y-%m-%d"),
            "status": "active",
            "goals": [""] * 10
        }
        data["participants"].append(participant)
    
    def get_user_goals(self, user_id: int, data: Dict) -> List[str]:
        """Получает цели пользователя (всегда возвращает список из 10 элементов)"""
        for participant in data["participants"]:
            if participant["user_id"] == user_id:
                goals = participant.get("goals", [])
                # Гарантируем, что всегда возвращаем список из 10 элементов
                if len(goals) < 10:
                    goals.extend([""] * (10 - len(goals)))
                return goals[:10]  # Обрезаем до 10 элементов, если больше
        # Если пользователь не найден, возвращаем пустой список из 10 элементов
        return [""] * 10
    
    def set_user_goal(self, user_id: int, goal_number: int, goal_text: str, data: Dict) -> None:
        """Устанавливает цель пользователя"""
        for participant in data["participants"]:
            if participant["user_id"] == user_id:
                if 1 <= goal_number <= 10:
                    participant["goals"][goal_number - 1] = goal_text
                break
    
    def save_daily_report(self, user_id: int, day: int, goals_progress: Dict[int, str], rest_day: bool, data: Dict) -> None:
        """Сохраняет ежедневный отчет"""
        # Проверяем, нет ли уже отчета за этот день
        for report in data["reports"]:
            if report["user_id"] == user_id and report["day"] == day:
                # Обновляем существующий отчет
                report["date"] = datetime.now().strftime("%Y-%m-%d")
                report["rest_day"] = rest_day
                for goal_num, progress in goals_progress.items():
                    if 1 <= goal_num <= 10:
                        report["progress"][goal_num - 1] = progress
                return
        
        # Создаем новый отчет
        progress_list = [""] * 10
        for goal_num, progress_text in goals_progress.items():
            if 1 <= goal_num <= 10:
                progress_list[goal_num - 1] = progress_text
        
        report = {
            "user_id": user_id,
            "day": day,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "progress": progress_list,
            "rest_day": rest_day
        }
        data["reports"].append(report)
    
    def get_user_reports_count(self, user_id: int, data: Dict) -> int:
        """Получает количество отчетов пользователя"""
        count = 0
        for report in data["reports"]:
            if report["user_id"] == user_id:
                count += 1
        return count
    
    async def get_current_day_async(self) -> int:
        """Получает текущий день из настроек или вычисляет"""
        try:
            settings = await self.get_settings()
            if settings.get("current_day"):
                return int(settings.get("current_day"))
        except:
            pass
        return self.get_current_day()
    
    def get_current_day(self, start_date: Optional[str] = None) -> int:
        """Вычисляет текущий день игры"""
        if start_date is None:
            # Старт игры - 5 ноября (из game_concept.txt)
            # Определяем год автоматически
            now = datetime.now()
            year = now.year
            # Если ноябрь еще не наступил, берем прошлый год
            if now.month < 11:
                year -= 1
            start_date = f"{year}-11-05"
        
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            now = datetime.now()
            delta = now - start
            day = delta.days + 1
            return max(1, min(day, 90))  # Ограничиваем от 1 до 90
        except Exception as e:
            logging.error(f"Ошибка вычисления дня: {e}")
            return 1
